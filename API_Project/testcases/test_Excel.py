from openpyxl import load_workbook

from API_Project.utils.mysql import execute_mysql

class TestExcel:

    def setup(self):
        # 获取excel
        self.wb = load_workbook('C:\\Users\\WBPC0154\\Desktop\\医疗项目2.xlsx')

    def test_getStatisTotal(self):

        # 获取sheet
        self.sheet = self.wb.get_sheet_by_name('收费项目人工抽检')

        for i in range(2, 41):
            # 获取单元格
            code = self.sheet.cell(row=i, column=1)
            print(code.value)

            sql = """SELECT std_code,std_name,std_score FROM project_安徽医疗网 WHERE code=%s"""
            std_value = execute_mysql(sql, params=[code.value, ])

            if std_value == ():
                continue
            else:
                print(std_value[0][1])
                code1 = self.sheet.cell(row=i, column=3)
                code1.value = std_value[0][0]

                code2 = self.sheet.cell(row=i, column=4)
                code2.value = std_value[0][1]

                code3 = self.sheet.cell(row=i, column=5)
                code3.value = std_value[0][2]

        self.wb.save(r'C:\\Users\\WBPC0154\\Desktop\\医疗项目2.xlsx')




